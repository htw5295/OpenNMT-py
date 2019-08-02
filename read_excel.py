import openpyxl
import codecs
import re
from tqdm import tqdm

wb = openpyxl.load_workbook('./data/navi.xlsx')

# 현재 Active Sheet 얻기
# ws = wb.get_sheet_by_name("Sheet1")

with codecs.open('./data/src-train.txt', 'w', encoding='utf-8') as src, codecs.open('./data/tgt-train.txt', 'w') as tgt:
    sheet_names = wb.sheetnames
    check = False
    count = 0
    regex = re.compile(r"<.*?>(.*?)</.*?>")
    for sheet_name in sheet_names:
        if sheet_name.startswith("|"):
            check = True
            continue
        else:
            if check and count < 8:
                sheet = wb.get_sheet_by_name(sheet_name)
                intent = ''
                total = sheet.max_row
                for i, row in tqdm(enumerate(sheet.rows), desc='read line', total=total):
                    if i == 0:
                        continue
                    elif i == 1:
                        intent = row[0].value
                    else:
                        m = regex.findall(row[4].value)
                        tgt_value = row[2].value
                        if len(m) == 0:
                            src.write(intent + '\n')
                        else:
                            src.write(intent + '\t' + '\t'.join(m) + '\n')
                            # for location in m:
                            #     tgt_value = location + ' ' + tgt_value
                            # tgt_value = tgt_value.replace('\t\t', '\t')
                        # if tgt_value[0] == '\t':
                        #     tgt_value = tgt_value[1:]
                        tgt.write(tgt_value + '\n')
                count += 1
